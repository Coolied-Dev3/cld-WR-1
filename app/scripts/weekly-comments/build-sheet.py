"""コメント案の確認用 Excel シートを作成する。

    python scripts/weekly-comments/build-sheet.py <reports.json> <out.xlsx> [drafts.json] [--sheet 2026-09-07]

reports.json は export-reports.mjs の出力。
drafts.json は {"<週報ID>": {"draft": "コメント案", "register": "はい|いいえ", "note": "備考"}} 形式(任意)。

**1回の作成につき1シート**。シート名は「コメントを作成した日」の日付(既定は本日、--sheet で指定可)。
対象週の日付ではないことに注意。出力先のブックが既にあればシートを追加する
(同じ日付のシートがある場合は作り直す)。シートは日付の新しい順に並べ、末尾に「使い方」を置く。

各シートの L列(コメント案)・M列(登録)・N列(備考) だけが編集対象。
行の追加・削除や A列(週報ID)の変更はしないこと(read-sheet.py が週報IDで突き合わせるため)。
"""

import json
import os
import sys
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Meiryo"

RATING = {"excellent": "◎ 期待以上", "good": "○ 期待どおり", "fair": "△ やや不足", "poor": "✕ 不足"}
ROLE = {"member": "メンバー", "manager": "所属長", "executive": "役員", "admin": "管理者"}
COMPLIANCE = {"none": "なし", "concern": "懸念あり", "issue": "問題あり"}

COLUMNS = [
    ("週報ID", 9),
    ("対象週", 12),
    ("事業室", 12),
    ("氏名", 14),
    ("役割", 10),
    ("自己評価", 12),
    ("業務内容", 44),
    ("自己評価コメント", 40),
    ("課題", 40),
    ("対策", 40),
    ("コンプラ", 11),
    ("コメント案(編集可)", 62),
    ("登録", 8),
    ("備考(編集可)", 26),
]
EDIT_COLS = {12, 13, 14}  # L, M, N

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
SKIP_FILL = PatternFill("solid", fgColor="F2F2F2")
ALERT_FILL = PatternFill("solid", fgColor="FCE4EC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def issue_text(issues, key_cat, key_comment):
    """課題(または対策)を「分類 / 記述」の形にまとめる。複数ある場合は番号を振る。"""
    blocks = []
    for n, i in enumerate(issues, 1):
        cat, comment = i.get(key_cat), i.get(key_comment)
        if not cat and not comment:
            continue
        head = f"{n}. " if len(issues) > 1 else ""
        parts = [f"{head}【{cat}】"] if cat else [f"{head}【分類なし】"]
        if comment:
            parts.append(comment.replace("\r\n", "\n"))
        blocks.append("\n".join(parts))
    return "\n\n".join(blocks)


def estimate_height(values, widths):
    """折り返し後の行数からおおよその行高を決める(上限あり)。"""
    lines = 1
    for value, width in zip(values, widths):
        if not value:
            continue
        n = 0
        for line in str(value).split("\n"):
            n += max(1, -(-len(line) * 2 // width))  # 全角換算でざっくり
        lines = max(lines, n)
    return min(max(30, lines * 15), 210)


def build(reports_path, out_path, drafts_path=None, sheet_name=None):
    data = json.load(open(reports_path, encoding="utf-8"))
    week, reports = data["week"], data["reports"]
    drafts = json.load(open(drafts_path, encoding="utf-8")) if drafts_path else {}
    # シート名はコメントを作成した日(対象週ではない)
    sheet_name = sheet_name or date.today().isoformat()

    # 既存ブックがあれば追記する(1回の作成につき1シート)
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        wb = Workbook()
        del wb[wb.sheetnames[0]]
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = f"週報コメント案  作成日: {sheet_name}  対象週: {week}  投稿者: 小野崎 康己(役員)"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A2"] = (
        "黄色の L・M・N 列だけを編集してください。"
        "M列「登録」が「はい」の行だけが DB に登録されます。"
        "行の追加・削除、A列(週報ID)の変更はしないでください。"
    )
    ws["A2"].font = Font(name=FONT, size=10, color="C00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    header_row = 4
    for c, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=c, value=title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[header_row].height = 28

    widths = [w for _, w in COLUMNS]
    row = header_row + 1
    for r in reports:
        d = drafts.get(r["reportId"], {})
        already = r.get("existingComments", 0) > 0
        note = d.get("note", "")
        if already and not note:
            note = f"既にコメントが{r['existingComments']}件あります"
        values = [
            int(r["reportId"]),
            r["week"],
            r["team"],
            r["name"],
            ROLE.get(r["role"], r["role"]),
            RATING.get(r["selfRating"], r["selfRating"]),
            (r["workSummary"] or "").replace("\r\n", "\n"),
            (r["freeComment"] or "").replace("\r\n", "\n"),
            issue_text(r["issues"], "issue", "issueComment"),
            issue_text(r["issues"], "countermeasure", "countermeasureComment"),
            COMPLIANCE.get((r["compliance"] or {}).get("level"), "なし"),
            d.get("draft", ""),
            d.get("register", "はい" if not already else "いいえ"),
            note,
        ]
        skip = values[12] != "はい"
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if c in (1, 2, 5, 6, 11, 13) else "left",
            )
            cell.border = BORDER
            if c in EDIT_COLS:
                cell.fill = EDIT_FILL
            elif skip:
                cell.fill = SKIP_FILL
        if (r["compliance"] or {}).get("level", "none") != "none":
            ws.cell(row=row, column=11).fill = ALERT_FILL
            ws.cell(row=row, column=11).font = Font(name=FONT, size=10, bold=True, color="C00000")
        ws.row_dimensions[row].height = estimate_height(values, widths)
        row += 1

    last = row - 1
    dv = DataValidation(type="list", formula1='"はい,いいえ"', allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"M{header_row + 1}:M{last}")

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:N{last}"

    add_readme(wb, sheet_name, week, len(reports))
    # 日付シートを新しい順に並べ、「使い方」を末尾に置く
    wb._sheets.sort(key=lambda s: ("", "") if s.title == "使い方" else ("0", s.title), reverse=True)
    wb.save(out_path)
    print(f"{len(reports)}件 -> {out_path} [シート: {sheet_name} / 対象週: {week}]")


def add_readme(wb, sheet_name, week, count):
    if "使い方" in wb.sheetnames:
        del wb["使い方"]
    ws = wb.create_sheet("使い方")
    lines = [
        ("週報コメント案シートの使い方", True),
        ("", False),
        ("シートは1回の作成につき1枚です(シート名 = コメントを作成した日の日付)。新しいシートが左に追加されます。", False),
        (f"最新シート: {sheet_name} / 対象週: {week} / 対象件数: {count}件 / 投稿者: 小野崎 康己(役員, user_id=13)", False),
        ("", False),
        ("1. いちばん左(最新の日付)のシートを開き、L列「コメント案」を確認・修正してください。", False),
        ("   文面はそのまま週報のコメントとして本人に表示されます。", False),
        ("2. 登録しない行は M列「登録」を「いいえ」に変更してください。", False),
        ("3. N列「備考」は社内メモです。DBには登録されません。", False),
        ("4. 確認が終わったらファイルを上書き保存し、登録スクリプトを実行してください。", False),
        ("", False),
        ("やってはいけないこと", True),
        ("・行の追加・削除、並べ替え(週報IDで突き合わせるため崩れます)", False),
        ("・A列「週報ID」の変更", False),
        ("・シート名(作成日)の変更", False),
        ("・L列以外の内容の書き換え(登録内容には反映されません)", False),
        ("", False),
        ("登録コマンド(app フォルダで実行)", True),
        (f'  python scripts/weekly-comments/read-sheet.py "<このファイル>" ../comments/sheet.json --sheet {sheet_name}', False),
        ("  node scripts/weekly-comments/apply-comments.mjs ../comments/sheet.json --dry-run", False),
        ("  node scripts/weekly-comments/apply-comments.mjs ../comments/sheet.json", False),
        ("", False),
        ("--sheet を省略すると、いちばん新しい日付のシートが対象になります。", False),
        ("--dry-run を付けると登録内容の確認だけを行い、DBは変更しません。", False),
        ("同じ週報に既にコメントがある場合は二重投稿を防ぐためスキップされます。", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT, size=12 if bold and i == 1 else 10, bold=bold)
    ws.column_dimensions["A"].width = 100


if __name__ == "__main__":
    argv = sys.argv[1:]
    sheet = argv[argv.index("--sheet") + 1] if "--sheet" in argv else None
    positional = [a for i, a in enumerate(argv) if not a.startswith("--") and (i == 0 or argv[i - 1] != "--sheet")]
    build(positional[0], positional[1], positional[2] if len(positional) > 2 else None, sheet)
