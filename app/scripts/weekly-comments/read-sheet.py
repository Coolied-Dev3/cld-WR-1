"""確認・修正済みのコメント案シートを読み取り、登録用 JSON に変換する。

    python scripts/weekly-comments/read-sheet.py <sheet.xlsx> <out.json> [--sheet 2026-08-31]

シート名はコメントを作成した日の日付。--sheet を省略した場合は
いちばん新しい日付のシートを読む。

A列(週報ID)・L列(コメント案)・M列(登録)・N列(備考) だけを読む。
M列が「はい」で、かつコメント案が空でない行だけを出力する。
"""

import json
import sys
from openpyxl import load_workbook

HEADER_ROW = 4
COL_ID, COL_NAME, COL_DRAFT, COL_REGISTER, COL_NOTE = 1, 4, 12, 13, 14


def pick_sheet(wb, sheet_name):
    """読み取るシートを選ぶ。省略時はいちばん新しい日付のシート。"""
    sheets = sorted(s for s in wb.sheetnames if s != "使い方")
    if not sheets:
        raise SystemExit("日付のシートが見つかりません。")
    if sheet_name and sheet_name not in sheets:
        raise SystemExit(f"シート {sheet_name} がありません。存在するシート: {', '.join(sheets)}")
    return wb[sheet_name or sheets[-1]]


def read(sheet_path, out_path, sheet_name=None):
    wb = load_workbook(sheet_path, data_only=True)
    ws = pick_sheet(wb, sheet_name)

    if ws.cell(HEADER_ROW, COL_ID).value != "週報ID" or ws.cell(HEADER_ROW, COL_DRAFT).value != "コメント案(編集可)":
        raise SystemExit("列の並びが想定と違います。行や列を追加・削除していないか確認してください。")

    items, skipped = [], []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        report_id = ws.cell(row, COL_ID).value
        if report_id is None:
            continue
        name = ws.cell(row, COL_NAME).value or ""
        register = str(ws.cell(row, COL_REGISTER).value or "").strip()
        content = (ws.cell(row, COL_DRAFT).value or "").strip()
        entry = {
            "reportId": str(int(report_id)),
            "name": name,
            "content": content,
            "note": ws.cell(row, COL_NOTE).value or "",
        }
        if register != "はい":
            skipped.append({**entry, "reason": f"登録={register or '(空欄)'}"})
        elif not content:
            skipped.append({**entry, "reason": "コメント案が空欄"})
        else:
            items.append(entry)

    payload = {"source": str(sheet_path), "sheet": ws.title, "items": items, "skipped": skipped}
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"シート {ws.title}: 登録対象 {len(items)}件 / 対象外 {len(skipped)}件 -> {out_path}")
    for s in skipped:
        print(f"  対象外: {s['reportId']} {s['name']} ({s['reason']})")


if __name__ == "__main__":
    argv = sys.argv[1:]
    sheet = argv[argv.index("--sheet") + 1] if "--sheet" in argv else None
    positional = [a for i, a in enumerate(argv) if not a.startswith("--") and (i == 0 or argv[i - 1] != "--sheet")]
    read(positional[0], positional[1], sheet)
